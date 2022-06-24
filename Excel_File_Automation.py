from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

file_path = "C:\\Users\\Hen Benoish\\Desktop\\Regression"
file_path_save = 'Converted_File'
file_type = "*.xlsx"
NVM_COLOR = "ffa200"
COLUMN_COLOR = {
    "NVM": "ffa200",
    "Result": ["C6EFCE", "FFC7CE"],
    "Environment": ["5cf7ff", "e75cff"],
    "Header color name": "dddddd"
}

fail_counter = 0
pass_counter = 0


def check_files_in_folder():
    # files = Path(file_path).glob(file_type)
    # print("--------------------------------------------------------------------------")
    # print(f"List of ({file_type}) files found in the path ({file_path}):")
    # print("--------------------------------------------------------------------------")
    # for num, file in enumerate(files):
    #     print(f"{num + 1}. {file}")
    files = Path(file_path).glob(file_type)
    return files


def get_document_sheet(doc):
    # print("--------------------------------------------------------------------------")
    # print(f"List of sheets:")
    # print("--------------------------------------------------------------------------")
    # for num, i in enumerate(doc):
    #     sheets = str(i).split('"')[1]
    #     print(f"{num + 1}. {sheets}")
    return doc.worksheets


def auto_col_fixed_with(sheet_fix):
    for col in sheet_fix.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet_fix.column_dimensions[column].width = adjusted_width


def load_document(doc):
    return load_workbook(doc)


def change_header_background_color(header):
    column_number = 1
    while header[get_column_letter(column_number) + str(1)].value is not None:
        header[get_column_letter(column_number) + str(1)].fill = PatternFill("solid",
                                                                             fgColor=COLUMN_COLOR["Header color name"])
        column_number += 1


def change_col_background_color_by_name(col, col_name, color):
    column_number = 1
    while col[get_column_letter(column_number) + str(1)].value is not None:
        if col[get_column_letter(column_number) + str(1)].value == col_name:
            current_cell = 2  # Skip the header cell
            while sheet[get_column_letter(column_number) + str(current_cell)].value is not None:
                sheet[get_column_letter(column_number) + str(current_cell)].fill = PatternFill("solid",
                                                                                               fgColor=color)
                current_cell += 1
        column_number += 1


ttt = 0


def pass_fail_mark(col, col_name, color):
    global fail_counter
    global pass_counter
    column_number = 1
    while col[get_column_letter(column_number) + str(1)].value is not None:
        if col[get_column_letter(column_number) + str(1)].value in col_name:
            current_cell = 2  # Skip the header cell
            while sheet[get_column_letter(column_number) + str(current_cell)].value is not None:
                if sheet[get_column_letter(column_number) + str(current_cell)].value == "PASS":
                    sheet[get_column_letter(column_number) + str(current_cell)].fill = PatternFill("solid",
                                                                                                   fgColor=color[0])
                    pass_counter += 1
                elif sheet[get_column_letter(column_number) + str(current_cell)].value == "FAIL":
                    sheet[get_column_letter(column_number) + str(current_cell)].fill = PatternFill("solid",
                                                                                                   fgColor=color[1])
                    fail_counter += 1
                current_cell += 1
        column_number += 1


def phy_type_mark(col, col_name, color):
    column_number = 1
    while col[get_column_letter(column_number) + str(1)].value is not None:
        if col[get_column_letter(column_number) + str(1)].value in col_name:
            current_cell = 2  # Skip the header cell
            index = 0
            first_value = sheet[get_column_letter(column_number) + str(current_cell)].value
            while sheet[get_column_letter(column_number) + str(current_cell)].value is not None:
                if sheet[get_column_letter(column_number) + str(current_cell)].value == first_value:
                    sheet[get_column_letter(column_number) + str(current_cell)].fill = PatternFill("solid",
                                                                                                   fgColor=color[
                                                                                                       index % 2])
                else:
                    first_value = sheet[get_column_letter(column_number) + str(current_cell)].value
                    index += 1
                    sheet[get_column_letter(column_number) + str(current_cell)].fill = PatternFill("solid",
                                                                                                   fgColor=color[
                                                                                                       index % 2])
                current_cell += 1
        column_number += 1


for document in check_files_in_folder():
    current_doc = load_document(document)
    for sheet in get_document_sheet(current_doc):
        auto_col_fixed_with(sheet)
        change_header_background_color(sheet)
        change_col_background_color_by_name(sheet, "NVM", COLUMN_COLOR["NVM"])
        pass_fail_mark(sheet, ["Result", "Run 1", "Run 2", "Run 3"], COLUMN_COLOR["Result"])
        phy_type_mark(sheet, "Environment", COLUMN_COLOR["Environment"])

    doc_name = str(document).split("\\")[5]
    current_doc.save(f'{file_path_save}/{doc_name}')
    print(f'Doc Name:       {doc_name}\nTotal fail:     {fail_counter}\nTotal pass:     {pass_counter}\n\n')
    fail_counter = 0
    pass_counter = 0
